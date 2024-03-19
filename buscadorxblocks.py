import os
import re
import openpyxl
from xml.etree import ElementTree as ET

# Función para extraer el título de un archivo XML
def extract_title(xml_file):
    with open(xml_file, 'r', encoding='utf-8') as file:
        first_line = file.readline()
        match = re.search(r'display_name="([^"]+)"', first_line)
        if match:
            return match.group(1)
    return None

# Función para buscar y contar la palabra clave
def find_keyword(xml_file, nombre_xblock_buscado):
    keyword_ids = []
    keyword_found = False
    with open(xml_file, 'r', encoding='utf-8') as file:
        for line in file:
            if nombre_xblock_buscado in line:
                match = re.search(r'url_name="([^"]+)"', line)
                if match:
                    keyword_ids.append(match.group(1))
                if not keyword_found:
                    keyword_found = True
    return keyword_ids

# Carpeta donde se encuentran los archivos XML
folder_path = './'  # Cambia esto a la ubicación de tu carpeta

# Definir la palabra clave a buscar
nombre_xblock_buscado = 'dialogsquestionsxblock'

# Crear un archivo XLSX para guardar los resultados
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = 'Resultados'
worksheet['A1'] = 'Nombre'
worksheet['B1'] = 'ID [' + nombre_xblock_buscado + "] encontrados:"

row = 2
max_title_length = 0  # Variable para almacenar la longitud del título más largo

# Recorre los archivos XML en la carpeta
for filename in os.listdir(folder_path):
    if filename.endswith('.xml'):
        xml_file = os.path.join(folder_path, filename)
        title = extract_title(xml_file)
        if title is not None:
            keyword_ids = find_keyword(xml_file, nombre_xblock_buscado)
            if keyword_ids:
                worksheet.cell(row=row, column=1, value=title)
                max_title_length = max(max_title_length, len(title))
                col = 2
                for keyword_id in keyword_ids:
                    worksheet.cell(row=row, column=col, value=keyword_id)
                    col += 1
                row += 1

# Establecer el ancho de la columna 'A' según el título más largo
worksheet.column_dimensions['A'].width = max_title_length

# Guardar el archivo XLSX
workbook.save('resultados.xlsx')
print("Resultados guardados en 'resultados.xlsx'")
